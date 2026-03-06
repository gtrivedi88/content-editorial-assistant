"""Build-time excerpt extraction from style guide source documents.

Reads style guide PDFs, Markdown files, and XLSX spreadsheets, then
extracts relevant excerpts and populates the excerpt fields in each
guide's YAML mapping file.

This script runs at build time only and is NOT deployed to the cluster.
The source documents (PDFs, XLSX) are also build-time only artifacts.

Source documents processed:
    1. IBM Style Guide PDF (ibm-style-documentation.pdf)
    2. Red Hat Supplementary Style Guide PDF
    3. Modular Documentation Reference Guide PDF
    4. Minimalism Basics PDF
    5. Official Red Hat Product Names List XLSX
    6. Getting Started with Accessibility for Writers MD

Usage:
    python scripts/extract_excerpts.py \\
        --ibm-pdf style_guides/ibm/ibm-style-documentation.pdf \\
        --output-dir style_guides/

Dependencies (build-time only, not in production requirements):
    - PyMuPDF (fitz) for PDF text extraction
    - openpyxl for XLSX parsing
    - PyYAML for YAML read/write
"""

import argparse
import logging
import re
import sys
from pathlib import Path
from typing import Any, Dict, List

import yaml

logger = logging.getLogger(__name__)

# Maximum characters per excerpt (100-200 tokens ≈ 400-800 chars)
_MAX_EXCERPT_CHARS = 800


def parse_args(argv: List[str]) -> argparse.Namespace:
    """Parse command-line arguments for excerpt extraction.

    Args:
        argv: Command-line argument list (typically sys.argv[1:]).

    Returns:
        Parsed argument namespace.
    """
    parser = argparse.ArgumentParser(
        description='Extract excerpts from style guide source documents into YAML mappings.',
    )
    parser.add_argument(
        '--ibm-pdf',
        type=Path,
        help='Path to IBM Style Guide PDF',
    )
    parser.add_argument(
        '--red-hat-pdf',
        type=Path,
        help='Path to Red Hat Supplementary Style Guide PDF',
    )
    parser.add_argument(
        '--modular-docs-pdf',
        type=Path,
        help='Path to Modular Documentation Reference Guide PDF',
    )
    parser.add_argument(
        '--minimalism-pdf',
        type=Path,
        help='Path to Minimalism Basics PDF',
    )
    parser.add_argument(
        '--product-names-xlsx',
        type=Path,
        help='Path to Official Red Hat Product Names XLSX',
    )
    parser.add_argument(
        '--accessibility-md',
        type=Path,
        help='Path to Getting Started with Accessibility for Writers MD',
    )
    parser.add_argument(
        '--output-dir',
        type=Path,
        default=Path('style_guides'),
        help='Output directory for updated YAML files (default: style_guides/)',
    )
    parser.add_argument(
        '--verbose',
        action='store_true',
        help='Enable verbose logging output',
    )
    return parser.parse_args(argv)


def extract_pdf_text(pdf_path: Path, page_numbers: List[int]) -> str:
    """Extract text from specific pages of a PDF document.

    Uses PyMuPDF (fitz) to read PDF pages and extract text content.
    Applies margin cropping (top/bottom 8 percent) to remove headers,
    footers, and page numbers.  Detects and removes repeated text
    (running headers, watermarks) that appears on 3+ pages.

    Args:
        pdf_path: Path to the PDF file.
        page_numbers: List of 1-based page numbers to extract.

    Returns:
        Combined text from the requested pages.
    """
    try:
        import fitz  # PyMuPDF — build-time only dependency
    except ImportError:
        logger.error(
            "PyMuPDF (fitz) is required for PDF extraction. "
            "Install with: pip install PyMuPDF"
        )
        return ""

    if not page_numbers:
        return ""

    doc = fitz.open(str(pdf_path))
    valid_pages = _validate_page_numbers(page_numbers, len(doc), pdf_path)

    repeated = _detect_repeated_blocks(doc, valid_pages)
    pages_text = _extract_page_texts(doc, valid_pages, repeated)

    doc.close()
    return "\n\n".join(pages_text)


def _validate_page_numbers(
    page_numbers: List[int], total_pages: int, pdf_path: Path,
) -> List[int]:
    """Filter page numbers to valid range, logging warnings for invalid ones.

    Args:
        page_numbers: 1-based page numbers requested.
        total_pages: Total pages in the PDF document.
        pdf_path: Path for logging context.

    Returns:
        List of valid 1-based page numbers.
    """
    valid: List[int] = []
    for pn in page_numbers:
        if pn < 1 or pn > total_pages:
            logger.warning("Page %d out of range for %s (%d pages)", pn, pdf_path, total_pages)
        else:
            valid.append(pn)
    return valid


def _get_cropped_clip(page: Any) -> Any:
    """Build a clipping rectangle with 8% top/bottom margin crop.

    Args:
        page: A PyMuPDF page object.

    Returns:
        A fitz.Rect clipping rectangle.
    """
    import fitz
    rect = page.rect
    margin_y = rect.height * 0.08
    return fitz.Rect(rect.x0, rect.y0 + margin_y, rect.x1, rect.y1 - margin_y)


def _detect_repeated_blocks(
    doc: Any, valid_pages: List[int],
) -> set:
    """First pass: count text blocks to detect headers/footers/watermarks.

    Args:
        doc: An open PyMuPDF document.
        valid_pages: 1-based page numbers to scan.

    Returns:
        Set of block-text prefixes that appear on 3+ pages.
    """
    block_counts: Dict[str, int] = {}
    for pn in valid_pages:
        page = doc[pn - 1]
        clip = _get_cropped_clip(page)
        for block in page.get_text("blocks", clip=clip, sort=True):
            key = block[4].strip()[:80]
            if key:
                block_counts[key] = block_counts.get(key, 0) + 1

    if len(valid_pages) < 3:
        return set()
    return {k for k, v in block_counts.items() if v >= 3}


def _extract_page_texts(
    doc: Any, valid_pages: List[int], repeated: set,
) -> List[str]:
    """Second pass: extract text from pages excluding repeated blocks.

    Args:
        doc: An open PyMuPDF document.
        valid_pages: 1-based page numbers to extract.
        repeated: Set of block-text prefixes to skip.

    Returns:
        List of per-page text strings.
    """
    pages_text: List[str] = []
    for pn in valid_pages:
        page = doc[pn - 1]
        clip = _get_cropped_clip(page)
        blocks = page.get_text("blocks", clip=clip, sort=True)
        page_text = [
            block[4].strip()
            for block in blocks
            if block[4].strip()[:80] not in repeated and len(block[4].strip()) > 5
        ]
        if page_text:
            pages_text.append("\n".join(page_text))
    return pages_text


def extract_xlsx_data(xlsx_path: Path) -> List[Dict[str, Any]]:
    """Extract product name data from an XLSX spreadsheet.

    Uses openpyxl to parse the Red Hat product names list and
    return structured data for each product entry.  Handles
    merged cells and blank spacer rows gracefully.

    Args:
        xlsx_path: Path to the XLSX file.

    Returns:
        List of product name entry dicts.
    """
    try:
        import openpyxl
    except ImportError:
        logger.error(
            "openpyxl is required for XLSX extraction. "
            "Install with: pip install openpyxl"
        )
        return []

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        logger.warning("No active worksheet found in %s", xlsx_path)
        wb.close()
        return []

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    entries: List[Dict[str, Any]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            entry = dict(zip(headers, row))
            # Skip blank spacer rows and rows with no product name
            if not entry.get("product_name"):
                continue
            entries.append(entry)
        except (TypeError, ValueError) as exc:
            # Merged cells or malformed rows — log and skip
            logger.debug("Skipping malformed XLSX row: %s", exc)
            continue

    wb.close()
    logger.info("Extracted %d product entries from %s", len(entries), xlsx_path)
    return entries


def extract_markdown_sections(md_path: Path) -> Dict[str, str]:
    """Extract sections from a Markdown document.

    Parses the Markdown file by heading structure (h1-h3) and returns
    a dict mapping section headings to their content.

    Args:
        md_path: Path to the Markdown file.

    Returns:
        Dict mapping heading text to section content.
    """
    text = md_path.read_text(encoding="utf-8")
    sections: Dict[str, str] = {}
    current_heading = None
    current_content: List[str] = []

    for line in text.splitlines():
        match = re.match(r'^(#{1,3})\s+(.+)', line)
        if match:
            if current_heading:
                sections[current_heading] = "\n".join(current_content).strip()
            current_heading = match.group(2).strip()
            current_content = []
        elif current_heading is not None:
            current_content.append(line)

    if current_heading:
        sections[current_heading] = "\n".join(current_content).strip()

    logger.info("Extracted %d sections from %s", len(sections), md_path)
    return sections


def update_yaml_excerpts(
    yaml_path: Path, excerpts: Dict[str, str]
) -> None:
    """Update excerpt fields in a YAML mapping file.

    Reads the existing YAML, merges in extracted excerpt text
    for each rule entry, and writes the updated YAML back.

    Args:
        yaml_path: Path to the YAML mapping file.
        excerpts: Dict mapping rule_id to excerpt text.
    """
    if not excerpts:
        logger.info("No excerpts to update for %s", yaml_path)
        return

    with open(yaml_path, 'r', encoding='utf-8') as f:
        data = yaml.safe_load(f) or {}

    updated = _apply_excerpts_to_data(data, excerpts)

    with open(yaml_path, 'w', encoding='utf-8') as f:
        yaml.dump(data, f, default_flow_style=False, allow_unicode=True, sort_keys=False)

    logger.info("Updated %d excerpt fields in %s", updated, yaml_path)


def _apply_excerpts_to_data(
    data: Dict[str, Any], excerpts: Dict[str, str],
) -> int:
    """Walk YAML data tree and inject excerpt text into matching rules.

    Args:
        data: Parsed YAML mapping dict (mutated in place).
        excerpts: Dict mapping rule_id to excerpt text.

    Returns:
        Number of excerpt fields updated.
    """
    _GUIDE_KEYS = ("ibm_style", "red_hat_ssg", "accessibility", "modular_docs")
    updated = 0

    for category_data in data.values():
        if not isinstance(category_data, dict):
            continue
        for rule_key, rule_data in category_data.items():
            if not isinstance(rule_data, dict):
                continue
            rule_id = rule_data.get("rule_id", rule_key)
            if rule_id not in excerpts:
                continue
            updated += _set_excerpt_on_rule(rule_data, excerpts[rule_id], _GUIDE_KEYS)

    return updated


def _set_excerpt_on_rule(
    rule_data: Dict[str, Any],
    excerpt_text: str,
    guide_keys: tuple,
) -> int:
    """Set the excerpt field on the first matching guide sub-dict.

    Args:
        rule_data: A single rule mapping dict (mutated in place).
        excerpt_text: The excerpt text to inject.
        guide_keys: Tuple of guide sub-dict keys to check.

    Returns:
        1 if updated, 0 otherwise.
    """
    for guide_key in guide_keys:
        if guide_key in rule_data:
            rule_data[guide_key]["excerpt"] = excerpt_text
            return 1
    return 0


def _extract_excerpts_from_pdf(
    pdf_path: Path,
    yaml_path: Path,
    guide_key: str,
) -> Dict[str, str]:
    """Extract excerpts for all rules in a YAML mapping that have page refs.

    Reads the YAML to find rules with page numbers under the specified
    guide key, extracts text from those pages, and returns a dict of
    rule_id → excerpt text.

    Args:
        pdf_path: Path to the source PDF.
        yaml_path: Path to the YAML mapping file.
        guide_key: Guide sub-dict key (e.g., 'ibm_style', 'red_hat_ssg').

    Returns:
        Dict mapping rule_id to extracted excerpt text.
    """
    with open(yaml_path, 'r', encoding='utf-8') as f:
        data = yaml.safe_load(f) or {}

    page_refs = _collect_page_refs(data, guide_key)
    excerpts = _extract_from_page_refs(pdf_path, page_refs)

    logger.info(
        "Extracted %d excerpts from %s for %s",
        len(excerpts), pdf_path.name, yaml_path.name,
    )
    return excerpts


def _collect_page_refs(
    data: Dict[str, Any], guide_key: str,
) -> List[tuple]:
    """Collect (rule_id, pages) pairs from YAML data for a given guide.

    Args:
        data: Parsed YAML mapping dict.
        guide_key: Guide sub-dict key to search.

    Returns:
        List of (rule_id, pages_list) tuples.
    """
    refs: List[tuple] = []
    for category_data in data.values():
        if not isinstance(category_data, dict):
            continue
        for rule_key, rule_data in category_data.items():
            if not isinstance(rule_data, dict):
                continue
            guide_data = rule_data.get(guide_key)
            if not isinstance(guide_data, dict):
                continue
            pages = guide_data.get("pages", [])
            if pages:
                refs.append((rule_data.get("rule_id", rule_key), pages))
    return refs


def _extract_from_page_refs(
    pdf_path: Path, page_refs: List[tuple],
) -> Dict[str, str]:
    """Extract and truncate text for each (rule_id, pages) pair.

    Args:
        pdf_path: Path to the source PDF.
        page_refs: List of (rule_id, pages_list) tuples.

    Returns:
        Dict mapping rule_id to truncated excerpt text.
    """
    excerpts: Dict[str, str] = {}
    for rule_id, pages in page_refs:
        raw_text = extract_pdf_text(pdf_path, pages)
        if raw_text:
            excerpts[rule_id] = _truncate_excerpt(raw_text)
    return excerpts


def _truncate_excerpt(text: str) -> str:
    """Truncate excerpt text to fit within budget, breaking at sentence boundary.

    Args:
        text: Raw extracted text.

    Returns:
        Truncated text (max _MAX_EXCERPT_CHARS).
    """
    if len(text) <= _MAX_EXCERPT_CHARS:
        return text

    # Try to break at a sentence boundary
    truncated = text[:_MAX_EXCERPT_CHARS]
    last_period = truncated.rfind('. ')
    if last_period > _MAX_EXCERPT_CHARS // 2:
        return truncated[:last_period + 1]

    # Fall back to word boundary
    last_space = truncated.rfind(' ')
    if last_space > 0:
        return truncated[:last_space] + "..."

    return truncated + "..."


def main() -> None:
    """Run the excerpt extraction pipeline.

    Parses arguments, reads source documents, extracts relevant
    excerpts, and updates the YAML mapping files.
    """
    args = parse_args(sys.argv[1:])

    log_level = logging.DEBUG if args.verbose else logging.INFO
    logging.basicConfig(
        level=log_level,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    )

    logger.info("Starting excerpt extraction")
    logger.info("Output directory: %s", args.output_dir)

    total = 0
    total += _process_pdf_source(args.ibm_pdf, args.output_dir, "ibm", "ibm_style")
    total += _process_pdf_source(args.red_hat_pdf, args.output_dir, "red_hat", "red_hat_ssg")
    total += _process_pdf_source(args.modular_docs_pdf, args.output_dir, "modular_docs", "modular_docs")
    _process_xlsx_source(args.product_names_xlsx, args.output_dir)
    total += _process_accessibility_source(args.accessibility_md, args.output_dir)

    logger.info("Excerpt extraction complete — %d total excerpts populated", total)


def _process_pdf_source(
    pdf_path: Any, output_dir: Path, subdir: str, guide_key: str,
) -> int:
    """Process a single PDF source document and update its YAML mapping.

    Args:
        pdf_path: Path to the PDF (may be None).
        output_dir: Base output directory for style guides.
        subdir: Subdirectory name (e.g., 'ibm', 'red_hat').
        guide_key: Guide sub-dict key in the YAML.

    Returns:
        Number of excerpts extracted.
    """
    if not pdf_path or not pdf_path.exists():
        logger.warning("PDF not provided or not found: %s", pdf_path)
        return 0

    yaml_name = f"{subdir.replace('-', '_')}_style_mapping.yaml"
    if subdir == "modular_docs":
        yaml_name = "modular_docs_mapping.yaml"
    elif subdir == "ibm":
        yaml_name = "ibm_style_mapping.yaml"
    elif subdir == "red_hat":
        yaml_name = "red_hat_style_mapping.yaml"

    yaml_path = output_dir / subdir / yaml_name
    if not yaml_path.exists():
        logger.warning("YAML mapping not found: %s", yaml_path)
        return 0

    logger.info("Processing %s: %s", subdir, pdf_path)
    excerpts = _extract_excerpts_from_pdf(pdf_path, yaml_path, guide_key)
    update_yaml_excerpts(yaml_path, excerpts)
    return len(excerpts)


def _process_xlsx_source(xlsx_path: Any, output_dir: Path) -> None:
    """Process product names XLSX and write merge candidates.

    Args:
        xlsx_path: Path to the XLSX file (may be None).
        output_dir: Base output directory.
    """
    if not xlsx_path or not xlsx_path.exists():
        logger.warning("Product names XLSX not provided or not found")
        return
    logger.info("Processing product names: %s", xlsx_path)
    entries = extract_xlsx_data(xlsx_path)
    if entries:
        _write_product_name_candidates(entries, output_dir)


def _process_accessibility_source(
    md_path: Any, output_dir: Path,
) -> int:
    """Process accessibility Markdown and update its YAML mapping.

    Args:
        md_path: Path to the Markdown file (may be None).
        output_dir: Base output directory.

    Returns:
        Number of excerpts extracted.
    """
    if not md_path or not md_path.exists():
        logger.warning("Accessibility MD not provided or not found")
        return 0

    logger.info("Processing accessibility guide: %s", md_path)
    sections = extract_markdown_sections(md_path)
    if not sections:
        return 0

    acc_yaml = output_dir / "accessibility" / "accessibility_mapping.yaml"
    if not acc_yaml.exists():
        logger.warning("Accessibility YAML not found: %s", acc_yaml)
        return 0

    excerpts = _map_accessibility_sections(sections)
    update_yaml_excerpts(acc_yaml, excerpts)
    return len(excerpts)


def _map_accessibility_sections(
    sections: Dict[str, str],
) -> Dict[str, str]:
    """Map accessibility guide sections to rule IDs.

    Args:
        sections: Dict mapping heading text to section content.

    Returns:
        Dict mapping rule_id to excerpt text.
    """
    # Heuristic mapping: section heading keywords → rule_id
    keyword_to_rule: Dict[str, str] = {
        "alt text": "accessibility",
        "alternative text": "accessibility",
        "color": "accessibility",
        "contrast": "accessibility",
        "screen reader": "accessibility",
        "heading": "accessibility",
        "link": "accessibility",
        "image": "accessibility",
        "table": "accessibility",
    }

    excerpts: Dict[str, str] = {}
    for heading, content in sections.items():
        heading_lower = heading.lower()
        for keyword, rule_id in keyword_to_rule.items():
            if keyword in heading_lower:
                existing = excerpts.get(rule_id, "")
                snippet = _truncate_excerpt(content)
                if existing:
                    combined = f"{existing}\n\n{snippet}"
                    excerpts[rule_id] = _truncate_excerpt(combined)
                else:
                    excerpts[rule_id] = snippet
                break

    logger.info("Mapped %d accessibility sections to rule excerpts", len(excerpts))
    return excerpts


def _write_product_name_candidates(
    entries: List[Dict[str, Any]],
    output_dir: Path,
) -> None:
    """Write product name merge candidates for review.

    Generates a YAML file with product names extracted from the XLSX
    that can be reviewed and merged into product_names_config.yaml.

    Args:
        entries: List of product name dicts from XLSX.
        output_dir: Output directory for the candidates file.
    """
    candidates_path = output_dir / "product_name_candidates.yaml"
    candidates: List[Dict[str, str]] = []

    for entry in entries:
        name = str(entry.get("product_name", "")).strip()
        if not name:
            continue
        candidate = {"product_name": name}
        # Include abbreviation if present
        abbrev = entry.get("abbreviation") or entry.get("short_name")
        if abbrev:
            candidate["abbreviation"] = str(abbrev).strip()
        candidates.append(candidate)

    with open(candidates_path, 'w', encoding='utf-8') as f:
        yaml.dump(
            {"product_name_candidates": candidates},
            f,
            default_flow_style=False,
            allow_unicode=True,
            sort_keys=False,
        )

    logger.info(
        "Wrote %d product name candidates to %s",
        len(candidates), candidates_path,
    )


if __name__ == '__main__':
    main()
